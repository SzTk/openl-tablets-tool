"""
OpenL Tablets Excel Writer

OpenLWorkbook モデルから OpenL Tablets 形式の Excel ファイルを生成する。

各テーブル種別の行レイアウト:

  SimpleDecisionTable:
    行1 : [None, タイトル]
    行2 : [None, method_signature]
    行3 : [None, "Rules", "SimpleDecisionTable", table_name]
    行4 : [None, "Condition"×n列, "Result"×m列, "備考"]  ← 役割ラベル
    行5 : [None, "ID", "列名 (型)"×n+m, "説明"]          ← ヘッダ
    行6+: [None, id, 条件値×n, 結果値×m, notes]           ← データ

  DataTable:
    行1 : [None, タイトル]
    行2 : [None, "Data", table_type, table_name]
    行3 : [None, "列名 (型)"×n]                           ← ヘッダ
    行4+: [None, 値×n]                                    ← データ

  SpreadsheetTable:
    行1 : [None, タイトル]
    行2 : [None, description]
    行3 : [None, "【入力パラメータ】"]
    行4+: [None, name, value, description]  (パラメータ)
    行n : [None, "【計算ステップ（OpenL式）】"]
    行n+: [None, label, value, unit]        (計算ステップ)
    末尾: [None, note]                      (備考行)
"""

from __future__ import annotations
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from .models import (
    OpenLWorkbook,
    SimpleDecisionTable,
    DataTable,
    SpreadsheetTable,
    AnyTable,
)


# ---------------------------------------------------------------------------
# ユーティリティ
# ---------------------------------------------------------------------------

def _col_header(name: str, col_type: str) -> str:
    return f"{name} ({col_type})"


def _append_row(ws: Worksheet, values: list) -> None:
    """ws に 1 行追記する。"""
    ws.append(values)


# ---------------------------------------------------------------------------
# テーブル種別ごとのライタ
# ---------------------------------------------------------------------------

def _write_simple_decision_table(ws: Worksheet, table: SimpleDecisionTable) -> None:
    prefix: list = [None] * table.start_col

    # 行1: シグネチャー（"SimpleRules ..." 単一セル）
    _append_row(ws, prefix + [table.method_signature])

    # 行2: カラムヘッダー（条件列名 + 結果列名）
    headers = [c.name for c in table.conditions] + [r.name for r in table.results]
    _append_row(ws, prefix + headers)

    # 行3+: データ行（None = 空セル、"=" 始まりはテキスト型で書き込む）
    for rule in table.rules:
        vals = [rule.conditions.get(c.name) for c in table.conditions]
        vals += [rule.results.get(r.name) for r in table.results]
        row_values = prefix + vals
        _append_row(ws, row_values)
        row_num = ws.max_row
        for col_idx, val in enumerate(row_values, start=1):
            if isinstance(val, str) and val.startswith("="):
                cell = ws.cell(row=row_num, column=col_idx)
                cell._value = val
                cell.data_type = "s"


def _write_data_table(ws: Worksheet, table: DataTable) -> None:
    prefix: list = [None] * table.start_col

    # Enum 型判定: columns が value 1 列だけなら Enum
    is_enum = len(table.columns) == 1 and table.columns[0].name == "value"

    # 宣言行: "Datatype {type}" または "Datatype {type} <{col_type}>"（ヘッダー行なし）
    if is_enum:
        decl = f"Datatype {table.table_type} <{table.columns[0].col_type}>"
    else:
        decl = f"Datatype {table.table_type}"
    _append_row(ws, prefix + [decl])

    # データ行
    for row in table.rows:
        vals = [row.data.get(c.name) for c in table.columns]
        _append_row(ws, prefix + vals)


def _write_spreadsheet_table(ws: Worksheet, table: SpreadsheetTable) -> tuple[int, int | None]:
    """
    SpreadsheetTable を書き込む。
    戻り値: (header_row_num, last_step_row_num | None)
      - 呼び出し側がヘッダー行・末尾行に Bold / Border を適用するために使う。
    """
    # 行1: 関数シグネチャー（OpenL "Spreadsheet" キーワード行）
    _append_row(ws, [None, table.description])

    # 行2: カラムヘッダー
    _append_row(ws, [None] + list(table.column_names))
    hdr_row = ws.max_row

    # 行3+: 計算ステップ（column_names の列数に合わせて書く）
    # "= ..." で始まる OpenL 式は Excel 数式に解釈されないようテキスト型で書く
    three_cols = len(table.column_names) >= 3
    for step in table.steps:
        row_values = [None, step.label, step.unit, step.value] if three_cols else [None, step.label, step.value]
        _append_row(ws, row_values)
        step_row = ws.max_row
        for col_idx, val in enumerate(row_values, start=1):
            if isinstance(val, str) and val.startswith("="):
                cell = ws.cell(row=step_row, column=col_idx)
                cell._value = val
                cell.data_type = "s"

    last_step_row = ws.max_row if table.steps else None

    # 備考
    for note in table.notes:
        _append_row(ws, [None, note])

    return hdr_row, last_step_row


_WRITER_MAP = {
    "SimpleDecisionTable": _write_simple_decision_table,
    "DataTable": _write_data_table,
    "SpreadsheetTable": _write_spreadsheet_table,
}


# ---------------------------------------------------------------------------
# メインライタ
# ---------------------------------------------------------------------------

class OpenLWriter:
    def write(self, workbook: OpenLWorkbook, path: str | Path) -> None:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # デフォルトシートを削除

        # SpreadsheetTable の構造スタイル適用対象行を収集
        # (sheet_name, row_num, 'header' | 'last_step')
        struct_targets: list[tuple[str, int, str]] = []

        # 同じ sheet_name を持つテーブルは同じシートに追記する
        for table in workbook.tables:
            if table.sheet_name in wb.sheetnames:
                ws = wb[table.sheet_name]
                ws.append([])  # テーブル間に空行
            else:
                ws = wb.create_sheet(title=table.sheet_name)

            writer_fn = _WRITER_MAP.get(table.table_kind)
            if writer_fn is None:
                raise ValueError(f"未対応のテーブル種別: {table.table_kind}")

            if table.table_kind == "SpreadsheetTable":
                hdr_row, last_row = writer_fn(ws, table)
                struct_targets.append((table.sheet_name, hdr_row, "header"))
                if last_row:
                    struct_targets.append((table.sheet_name, last_row, "last_step"))
            else:
                writer_fn(ws, table)

        # SpreadsheetTable のヘッダー行・末尾行に Bold / Border を適用
        for sheet_name, row_num, style_type in struct_targets:
            ws = wb[sheet_name]
            for col in (2, 3, 4):
                cell = ws.cell(row=row_num, column=col)
                # フォント名・サイズは既存の値を引き継ぎ、Bold だけ上書き
                cell.font = Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=True,
                    italic=cell.font.italic,
                )
                if style_type == "last_step":
                    cell.border = Border(
                        top=Side(style="thin"),
                        bottom=cell.border.bottom,
                        left=cell.border.left,
                        right=cell.border.right,
                    )

        wb.save(str(path))
