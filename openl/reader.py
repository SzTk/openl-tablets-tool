"""
OpenL Tablets Excel Reader

Excel シートを解析し OpenLWorkbook モデルに変換する。

テーブル種別の判定: シート内のセルを走査し、OpenL キーワードで始まるセルを検出する。

  'SimpleRules ...'  → SimpleDecisionTable
  'Spreadsheet ...'  → SpreadsheetTable
  'Datatype ...'     → DataTable

1シートに複数テーブルが存在する場合も対応。
"""

from __future__ import annotations
from pathlib import Path
from typing import Any
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .models import (
    CellStyle,
    SheetDimensions,
    _DEFAULT_FONT_NAME,
    _DEFAULT_FONT_SIZE,
    ColumnDef,
    Rule,
    SimpleDecisionTable,
    DataTable,
    DataTableRow,
    SpreadsheetTable,
    SpreadsheetStep,
    OpenLWorkbook,
)


def _rows_as_lists(ws: Worksheet) -> list[list[Any]]:
    return [[cell.value for cell in row] for row in ws.iter_rows()]


# ---------------------------------------------------------------------------
# テーブルヘッダー検索
# ---------------------------------------------------------------------------

_KEYWORDS = ("SimpleRules", "Spreadsheet", "Datatype")


def _scan_table_headers(rows: list[list[Any]]) -> list[tuple[int, int, str]]:
    """
    シート内の全テーブル開始位置を返す: [(row_idx, col_idx, keyword), ...]
    キーワード + スペースで始まるセルを探す。
    """
    results: list[tuple[int, int, str]] = []
    seen_rows: set[int] = set()

    for row_idx, row in enumerate(rows):
        if row_idx in seen_rows:
            continue
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            s = str(cell).strip()
            for kw in _KEYWORDS:
                if s.startswith(kw + " "):
                    results.append((row_idx, col_idx, kw))
                    seen_rows.add(row_idx)
                    break
            else:
                continue
            break

    return results


# ---------------------------------------------------------------------------
# パーサ
# ---------------------------------------------------------------------------

def _parse_simple_rules(
    sheet_name: str,
    rows: list[list[Any]],
    start_row: int,
    start_col: int,
    end_row: int,
) -> SimpleDecisionTable:
    """
    SimpleRules テーブルを解析する。

    形式:
      rows[start_row][start_col] : "SimpleRules ReturnType MethodName(params)"
      rows[start_row+1][start_col:]: 列名（条件列...最後が結果列）
      rows[start_row+2:end_row]  : データ行
    """
    method_sig = str(rows[start_row][start_col] or "").strip()

    # テーブル名をシグネチャから抽出
    # "SimpleRules TheftRating VehicleTheftRating (BodyType bodyType, ...)" → "VehicleTheftRating"
    parts = method_sig.split()
    table_name = parts[2].split("(")[0].strip() if len(parts) >= 3 else method_sig

    # 列ヘッダー行
    header_row = rows[start_row + 1] if start_row + 1 < len(rows) else []
    col_headers: list[tuple[int, str]] = [
        (col_idx, str(val).strip())
        for col_idx in range(start_col, len(header_row))
        if (val := header_row[col_idx]) is not None
    ]

    if not col_headers:
        return SimpleDecisionTable(
            sheet_name=sheet_name, title="", method_signature=method_sig,
            table_name=table_name, conditions=[], results=[], rules=[],
        )

    # 最後の列が結果列、それ以外は条件列
    conditions = [
        ColumnDef(name=name, col_type="String", role="condition")
        for _, name in col_headers[:-1]
    ]
    result_cols = [ColumnDef(name=col_headers[-1][1], col_type="String", role="result")]
    col_indices = [idx for idx, _ in col_headers]
    cond_names = [name for _, name in col_headers[:-1]]
    result_name = col_headers[-1][1]

    # データ行
    rules: list[Rule] = []
    rule_id = 1
    for row in rows[start_row + 2:end_row]:
        if all(v is None for v in row):
            continue
        vals = [row[idx] if idx < len(row) else None for idx in col_indices]
        if all(v is None for v in vals):
            continue
        rules.append(Rule(
            id=rule_id,
            conditions={name: vals[i] for i, name in enumerate(cond_names)},
            results={result_name: vals[-1]},
        ))
        rule_id += 1

    return SimpleDecisionTable(
        sheet_name=sheet_name,
        title="",
        method_signature=method_sig,
        table_name=table_name,
        conditions=conditions,
        results=result_cols,
        rules=rules,
        start_col=start_col,
    )


def _parse_spreadsheet(
    sheet_name: str,
    rows: list[list[Any]],
    start_row: int,
    start_col: int,
    end_row: int,
) -> SpreadsheetTable:
    """
    Spreadsheet テーブルを解析する。

    形式:
      rows[start_row][start_col]   : "Spreadsheet ReturnType MethodName(params)"
      rows[start_row+1][start_col:]: "Step", "Description", "Value" ヘッダー
      rows[start_row+2:end_row]    : ステップ行 (label, description, value)
    """
    method_sig = str(rows[start_row][start_col] or "").strip()

    # ヘッダー行からカラム名を読み取る（None セルで打ち切り、最低1列）
    hdr_row = rows[start_row + 1] if start_row + 1 < len(rows) else []
    col_names: list[str] = []
    for i in range(3):
        v = hdr_row[start_col + i] if start_col + i < len(hdr_row) else None
        if v is None:
            break
        col_names.append(str(v).strip())
    if not col_names:
        col_names = ["Step", "Description", "Value"]

    three_cols = len(col_names) >= 3
    steps: list[SpreadsheetStep] = []
    for row in rows[start_row + 2:end_row]:
        if all(v is None for v in row):
            continue
        label = row[start_col] if start_col < len(row) else None
        if three_cols:
            desc = row[start_col + 1] if start_col + 1 < len(row) else None
            val  = row[start_col + 2] if start_col + 2 < len(row) else None
        else:
            desc = None
            val  = row[start_col + 1] if start_col + 1 < len(row) else None
        if label is None:
            continue
        steps.append(SpreadsheetStep(
            label=str(label).strip(),
            value=val,
            unit=str(desc).strip() if desc is not None else None,
        ))

    return SpreadsheetTable(
        sheet_name=sheet_name,
        title="",
        description=method_sig,
        steps=steps,
        column_names=col_names,
        start_col=start_col,
    )


def _parse_datatype(
    sheet_name: str,
    rows: list[list[Any]],
    start_row: int,
    start_col: int,
    end_row: int,
) -> DataTable:
    """
    Datatype テーブルを解析する。

    形式:
      rows[start_row][start_col]: "Datatype TypeName" または "Datatype TypeName <BaseType>"
      rows[start_row+1:end_row] : フィールド行 (fieldType, fieldName, defaultValue)
                                   または enum 値行 (value のみ)
    """
    decl = str(rows[start_row][start_col] or "").strip()
    parts = decl.split()
    type_name = parts[1].strip("<>") if len(parts) >= 2 else decl
    is_enum = "<" in decl

    if is_enum:
        columns = [ColumnDef(name="value", col_type="String", role="data")]
    else:
        columns = [
            ColumnDef(name="fieldType",     col_type="String", role="data"),
            ColumnDef(name="fieldName",     col_type="String", role="data"),
            ColumnDef(name="defaultValue",  col_type="String", role="data"),
        ]

    data_rows: list[DataTableRow] = []
    for row in rows[start_row + 1:end_row]:
        if all(v is None for v in row):
            continue
        v1 = row[start_col]     if start_col     < len(row) else None
        v2 = row[start_col + 1] if start_col + 1 < len(row) else None
        v3 = row[start_col + 2] if start_col + 2 < len(row) else None
        if v1 is None and v2 is None:
            continue
        if is_enum:
            data_rows.append(DataTableRow(data={"value": v1}))
        else:
            data_rows.append(DataTableRow(data={"fieldType": v1, "fieldName": v2, "defaultValue": v3}))

    return DataTable(
        sheet_name=sheet_name,
        title="",
        table_type=type_name,
        table_name=type_name,
        columns=columns,
        rows=data_rows,
        start_col=start_col,
    )


_PARSERS = {
    "SimpleRules": _parse_simple_rules,
    "Spreadsheet": _parse_spreadsheet,
    "Datatype":    _parse_datatype,
}


# ---------------------------------------------------------------------------
# スタイル・サイズ読み取り
# ---------------------------------------------------------------------------

_DEFAULT_FONT_COLORS = {"FF000000", "00000000", ""}
_DEFAULT_FILL_COLORS = {"00000000", "FFFFFFFF", ""}


def _border_style(side) -> str | None:
    return side.style if side and side.style else None


def _read_sheet_styles(ws: Worksheet) -> dict[str, CellStyle]:
    styles: dict[str, CellStyle] = {}
    for row in ws.iter_rows():
        for cell in row:
            font = cell.font
            fill = cell.fill
            border = cell.border
            nf = cell.number_format or "General"

            font_name: str | None = None
            if font.name and font.name != _DEFAULT_FONT_NAME:
                font_name = font.name

            font_size: float | None = None
            if font.size and font.size != _DEFAULT_FONT_SIZE:
                font_size = float(font.size)

            font_color: str | None = None
            if font.color and font.color.type == "rgb":
                rgb = font.color.rgb
                if rgb not in _DEFAULT_FONT_COLORS:
                    font_color = rgb

            fill_color: str | None = None
            if fill.fill_type == "solid" and fill.fgColor.type == "rgb":
                rgb = fill.fgColor.rgb
                if rgb not in _DEFAULT_FILL_COLORS:
                    fill_color = rgb

            style = CellStyle(
                bold=bool(font.bold),
                italic=bool(font.italic),
                font_name=font_name,
                font_size=font_size,
                font_color=font_color,
                fill_color=fill_color,
                number_format=nf,
                border_top=_border_style(border.top),
                border_bottom=_border_style(border.bottom),
                border_left=_border_style(border.left),
                border_right=_border_style(border.right),
            )
            if not style.is_default():
                styles[cell.coordinate] = style
    return styles


def _read_sheet_dimensions(ws: Worksheet) -> SheetDimensions | None:
    col_widths = {ltr: dim.width for ltr, dim in ws.column_dimensions.items() if dim.width}
    row_heights = {str(n): dim.height for n, dim in ws.row_dimensions.items() if dim.height}
    if not col_widths and not row_heights:
        return None
    return SheetDimensions(column_widths=col_widths, row_heights=row_heights)


# ---------------------------------------------------------------------------
# メインリーダ
# ---------------------------------------------------------------------------

class OpenLReader:
    def read(self, path: str | Path) -> OpenLWorkbook:
        wb = openpyxl.load_workbook(str(path), data_only=True)
        workbook = OpenLWorkbook(source_file=str(Path(path).name))

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = _rows_as_lists(ws)

            # 空シートはスキップ
            if all(all(v is None for v in row) for row in rows):
                continue

            headers = _scan_table_headers(rows)

            if not headers:
                continue  # OpenL テーブルが見つからないシートはスキップ

            for i, (row_idx, col_idx, keyword) in enumerate(headers):
                end_row = headers[i + 1][0] if i + 1 < len(headers) else len(rows)
                parser = _PARSERS[keyword]
                table = parser(sheet_name, rows, row_idx, col_idx, end_row)
                workbook.tables.append(table)

            sheet_style = _read_sheet_styles(ws)
            if sheet_style:
                workbook.sheet_styles[sheet_name] = sheet_style

            dims = _read_sheet_dimensions(ws)
            if dims:
                workbook.sheet_dimensions[sheet_name] = dims

        return workbook
