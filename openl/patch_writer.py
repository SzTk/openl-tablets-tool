"""
OpenL Tablets In-Place Patch Writer

編集後の OpenLWorkbook と元の Excel ファイルを比較し、変更箇所だけを
元ファイルに適用して書き戻す。編集対象外セルの書式・レイアウトは
そのまま保持される。
"""

from __future__ import annotations
from copy import copy
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .models import AnyTable, OpenLWorkbook
from .reader import ParsedTable, read_with_positions
from .writer import (
    _write_simple_decision_table,
    _write_data_table,
    _write_spreadsheet_table,
    _style_spreadsheet_struct_row,
)


TableIdentity = tuple[str, str, str]

_WRITER_MAP = {
    "SimpleDecisionTable": _write_simple_decision_table,
    "DataTable": _write_data_table,
    "SpreadsheetTable": _write_spreadsheet_table,
}


def _table_identity(table: AnyTable) -> TableIdentity:
    """(sheet_name, table_kind, identifier) でテーブルを一意に識別する。

    SpreadsheetTable は table_name を持たないため、description（シグネチャ）
    の3単語目から '(' より前を抽出する。SimpleRules/Spreadsheet 共通の
    シグネチャ書式 "<Keyword> <ReturnType> <Name>(<params>)" に依存する。
    """
    if table.table_kind == "SpreadsheetTable":
        parts = (table.description or "").split()
        ident = parts[2].split("(")[0] if len(parts) >= 3 else (table.description or "")
        return (table.sheet_name, table.table_kind, ident.strip())
    return (table.sheet_name, table.table_kind, table.table_name)


def _structure_key(table: AnyTable) -> tuple:
    """列構成の比較用キー。before/afterで異なれば delete+recreate の対象。"""
    if table.table_kind == "SimpleDecisionTable":
        return (
            tuple((c.name, c.col_type) for c in table.conditions),
            tuple((c.name, c.col_type) for c in table.results),
        )
    if table.table_kind == "DataTable":
        return tuple((c.name, c.col_type) for c in table.columns)
    return tuple(table.column_names)  # SpreadsheetTable


def _comparison_tuples(table: AnyTable) -> list[tuple]:
    """diff用の比較タプル列。ParsedTable.item_rows と要素数・順序が対応する。"""
    if table.table_kind == "SimpleDecisionTable":
        return [
            tuple(rule.conditions.get(c.name) for c in table.conditions)
            + tuple(rule.results.get(r.name) for r in table.results)
            + (rule.notes,)
            for rule in table.rules
        ]
    if table.table_kind == "DataTable":
        return [tuple(row.data.get(c.name) for c in table.columns) for row in table.rows]
    return [(step.label, step.value, step.unit) for step in table.steps]


def _row_cell_values(table: AnyTable, index: int) -> list[Any]:
    """item index 番目のデータ行を、start_col直後から並ぶセル値のリストとして返す。"""
    if table.table_kind == "SimpleDecisionTable":
        rule = table.rules[index]
        return [rule.conditions.get(c.name) for c in table.conditions] + \
               [rule.results.get(r.name) for r in table.results]
    if table.table_kind == "DataTable":
        row = table.rows[index]
        return [row.data.get(c.name) for c in table.columns]
    step = table.steps[index]
    if len(table.column_names) >= 3:
        return [step.label, step.unit, step.value]
    return [step.label, step.value]


def _set_cell(ws: Worksheet, row: int, col: int, value: Any) -> None:
    """セルに値を書き込む。'=' 始まりの文字列は OpenL 式としてテキスト型で書く。"""
    cell = ws.cell(row=row, column=col)
    if isinstance(value, str) and value.startswith("="):
        cell._value = value
        cell.data_type = "s"
    else:
        cell.value = value


def _copy_row_style(ws: Worksheet, template_row: int, target_row: int, columns: range) -> None:
    """target_row の各セルに template_row のスタイルをコピーする。"""
    for col in columns:
        src = ws.cell(row=template_row, column=col)
        dst = ws.cell(row=target_row, column=col)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format


def _insert_table_rows(
    ws: Worksheet, after: AnyTable, j1: int, j2: int, anchor_row: int, start_col: int,
) -> None:
    """anchor_row の直後に after の item j1..j2 を挿入し、anchor_row のスタイルをコピーする。"""
    count = j2 - j1
    insert_at = anchor_row + 1
    ws.insert_rows(insert_at, count)
    n_cols = len(_row_cell_values(after, j1))
    columns = range(start_col, start_col + n_cols)
    for offset in range(count):
        row_num = insert_at + offset
        _copy_row_style(ws, anchor_row, row_num, columns)
        for col, val in zip(columns, _row_cell_values(after, j1 + offset)):
            _set_cell(ws, row_num, col, val)


def _patch_table_rows(ws: Worksheet, before: ParsedTable, after: AnyTable) -> None:
    """before/after の比較タプルを SequenceMatcher で diff し、行単位でパッチを適用する。

    行番号がずれないよう、opcode はシート内で行番号の大きい方から処理する
    （SequenceMatcher.get_opcodes() の逆順）。item_rows はテーブル内で
    連続した行番号であることを前提とする。
    """
    before_tuples = _comparison_tuples(before.table)
    after_tuples = _comparison_tuples(after)
    matcher = SequenceMatcher(a=before_tuples, b=after_tuples, autojunk=False)
    start_col = after.start_col + 1  # 1-based

    for tag, i1, i2, j1, j2 in reversed(matcher.get_opcodes()):
        if tag == "equal":
            continue

        if tag == "replace":
            common = min(i2 - i1, j2 - j1)
            for k in range(common):
                row_num = before.item_rows[i1 + k]
                for col, val in zip(
                    range(start_col, start_col + len(_row_cell_values(after, j1 + k))),
                    _row_cell_values(after, j1 + k),
                ):
                    _set_cell(ws, row_num, col, val)
            if i2 - i1 > common:
                rows = before.item_rows[i1 + common:i2]
                ws.delete_rows(rows[0], len(rows))
            elif j2 - j1 > common:
                anchor = before.item_rows[i1 + common - 1]
                _insert_table_rows(ws, after, j1 + common, j2, anchor, start_col)

        elif tag == "delete":
            rows = before.item_rows[i1:i2]
            ws.delete_rows(rows[0], len(rows))

        elif tag == "insert":
            if i1 > 0:
                anchor = before.item_rows[i1 - 1]
            elif before.item_rows:
                anchor = before.item_rows[0] - 1
            else:
                anchor = before.start_row + before.header_rows - 1
            _insert_table_rows(ws, after, j1, j2, anchor, start_col)


def _delete_table(ws: Worksheet, before: ParsedTable) -> None:
    """before側のみに存在するテーブルの行範囲（シグネチャ行〜区切り空行まで）を削除する。"""
    n = before.end_row - before.start_row + 1
    ws.delete_rows(before.start_row, n)


def _append_table(ws: Worksheet, table: AnyTable, *, separator: bool) -> None:
    """afterのみに存在するテーブルをシート末尾に追記する。既存シートなら空行を挟む。"""
    has_content = any(c.value is not None for row in ws.iter_rows() for c in row)
    if separator and has_content:
        ws.append([])
    if table.table_kind == "SpreadsheetTable":
        hdr_row, last_row = _write_spreadsheet_table(ws, table)
        _style_spreadsheet_struct_row(ws, hdr_row, "header", table.start_col)
        if last_row:
            _style_spreadsheet_struct_row(ws, last_row, "last_step", table.start_col)
    else:
        _WRITER_MAP[table.table_kind](ws, table)


def _recreate_table(ws: Worksheet, before: ParsedTable, after: AnyTable) -> None:
    """列構成が変わったテーブルを削除して同じ位置に再作成する。

    このテーブルのみ書式が openpyxl デフォルトに戻る（他のテーブル・セルは無影響）。
    削除/挿入の対象はテーブル本体（header_rows + item_rows）のみとし、
    end_row に含まれる末尾の区切り空行・備考行はそのまま残す。
    """
    n_before = before.header_rows + len(before.item_rows)
    ws.delete_rows(before.start_row, n_before)

    tmp_ws = openpyxl.Workbook().active
    _WRITER_MAP[after.table_kind](tmp_ws, after)
    n_after = tmp_ws.max_row

    ws.insert_rows(before.start_row, n_after)
    for r in range(1, n_after + 1):
        for c in range(1, tmp_ws.max_column + 1):
            _set_cell(ws, before.start_row + r - 1, c, tmp_ws.cell(row=r, column=c).value)


def patch_write(edited: OpenLWorkbook, original_path: str | Path, out_path: str | Path) -> None:
    """編集後の edited を、original_path の書式・レイアウトを保ったまま out_path に書き出す。"""
    wb = openpyxl.load_workbook(str(original_path))
    before_parsed = read_with_positions(original_path)

    after_by_id = {_table_identity(t): t for t in edited.tables}

    original_sheets = set(wb.sheetnames)
    sheet_names = {p.table.sheet_name for p in before_parsed} | \
        {t.sheet_name for t in edited.tables}

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(title=sheet_name)
        ws = wb[sheet_name]

        # 行の挿入/削除によるズレを避けるため、行番号の大きい方から処理する
        sheet_before = sorted(
            (p for p in before_parsed if p.table.sheet_name == sheet_name),
            key=lambda p: p.start_row,
            reverse=True,
        )
        handled: set[TableIdentity] = set()

        for parsed in sheet_before:
            ident = _table_identity(parsed.table)
            handled.add(ident)
            after = after_by_id.get(ident)
            if after is None:
                _delete_table(ws, parsed)
            elif _structure_key(parsed.table) != _structure_key(after):
                _recreate_table(ws, parsed, after)
            else:
                _patch_table_rows(ws, parsed, after)

        appended_any = False
        for ident, table in after_by_id.items():
            if ident[0] == sheet_name and ident not in handled:
                _append_table(ws, table, separator=sheet_name in original_sheets or appended_any)
                appended_any = True

    wb.save(str(out_path))
