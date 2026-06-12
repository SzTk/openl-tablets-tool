from copy import copy
from pathlib import Path

import openpyxl

from openl.reader import OpenLReader
from openl.patch_writer import (
    _table_identity,
    _structure_key,
    _comparison_tuples,
    _row_cell_values,
    _set_cell,
)
from openl.models import Rule, DataTableRow, SpreadsheetStep, SimpleDecisionTable, SpreadsheetTable, ColumnDef

SHOP_POLICY = Path(__file__).parent.parent / "examples" / "ShopPolicy.xlsx"
AUTO_POLICY = Path(__file__).parent.parent / "examples" / "AutoPolicyCalculation.xlsx"


def _shop_policy():
    return OpenLReader().read(SHOP_POLICY)


def _fonts_equal(font_a, font_b) -> bool:
    """cell.font (StyleProxy) 同士は __eq__ の非対称性のため常に False になるため、
    copy() で素の Font を取り出して比較する。"""
    return copy(font_a) == copy(font_b)


def test_table_identity_simple_decision_table():
    wb = _shop_policy()
    table = wb.get_table("FreeShipping")
    assert _table_identity(table) == ("FreeShipping", "SimpleDecisionTable", "IsFreeShipping")


def test_table_identity_spreadsheet_table():
    wb = _shop_policy()
    table = wb.get_table("Calculation")
    assert table.table_kind == "SpreadsheetTable"
    assert _table_identity(table) == ("Calculation", "SpreadsheetTable", "DetermineShopPolicy")


def test_structure_key_matches_for_unchanged_table():
    wb = _shop_policy()
    table = wb.get_table("FreeShipping")
    assert _structure_key(table) == _structure_key(table.model_copy(deep=True))


def test_structure_key_differs_when_condition_added():
    wb = _shop_policy()
    table = wb.get_table("FreeShipping")
    changed = table.model_copy(deep=True)
    from openl.models import ColumnDef
    changed.conditions.append(ColumnDef(name="新条件", col_type="String", role="condition"))
    assert _structure_key(table) != _structure_key(changed)


def test_comparison_tuples_simple_decision_table():
    wb = _shop_policy()
    table = wb.get_table("FreeShipping")
    tuples = _comparison_tuples(table)
    assert len(tuples) == 9
    assert tuples[0] == ("非会員", "< 3000", False, None)


def test_row_cell_values_simple_decision_table():
    wb = _shop_policy()
    table = wb.get_table("FreeShipping")
    assert _row_cell_values(table, 0) == ["非会員", "< 3000", False]


def test_comparison_tuples_spreadsheet_table():
    wb = _shop_policy()
    table = wb.get_table("Calculation")
    tuples = _comparison_tuples(table)
    assert tuples[0] == ("FreeShipping", "= IsFreeShipping(memberType, purchaseAmount)", "送料無料フラグ")


def test_set_cell_writes_formula_text_as_string():
    wb = openpyxl.Workbook()
    ws = wb.active
    _set_cell(ws, 1, 1, "= 1 + 1")
    cell = ws.cell(row=1, column=1)
    assert cell.value == "= 1 + 1"
    assert cell.data_type == "s"


def test_set_cell_writes_plain_value():
    wb = openpyxl.Workbook()
    ws = wb.active
    _set_cell(ws, 1, 1, 123)
    assert ws.cell(row=1, column=1).value == 123


def _all_cell_values(wb):
    """openpyxl Workbook の全シート・全セル値を {sheet_name: [[...], ...]} で返す。

    各行末尾の None セルは比較対象外とする。openpyxl は保存時にシートの
    dimensions を実際の内容に合わせて再計算するため、編集前後で末尾の
    空セル数が変わることがある（実際の値には影響しない）。
    """
    result = {}
    for name in wb.sheetnames:
        rows = []
        for row in wb[name].iter_rows():
            values = [c.value for c in row]
            while values and values[-1] is None:
                values.pop()
            rows.append(values)
        result[name] = rows
    return result


def test_patch_write_value_change_preserves_everything_else(tmp_path):
    from openl.patch_writer import patch_write

    edited = OpenLReader().read(SHOP_POLICY)
    table = edited.get_table("FreeShipping")
    table.rules[0].results["送料無料"] = True  # was False

    out_path = tmp_path / "out.xlsx"
    patch_write(edited, SHOP_POLICY, out_path)

    before_wb = openpyxl.load_workbook(SHOP_POLICY)
    after_wb = openpyxl.load_workbook(out_path)

    before = _all_cell_values(before_wb)
    after = _all_cell_values(after_wb)

    # 変更したセル（FreeShipping シート 3行目 D列 = 結果列）
    assert before["FreeShipping"][2][3] is False
    assert after["FreeShipping"][2][3] is True

    # 同じ行の他のセルは不変
    assert after["FreeShipping"][2][1:3] == before["FreeShipping"][2][1:3]

    # 変更したセル以外はすべて一致
    after["FreeShipping"][2][3] = before["FreeShipping"][2][3]
    assert after == before


def test_patch_write_append_rule_to_end(tmp_path):
    from openl.patch_writer import patch_write

    edited = OpenLReader().read(SHOP_POLICY)
    table = edited.get_table("FreeShipping")
    table.rules.append(Rule(
        id=10,
        conditions={"会員種別": "プレミアム会員", "購入金額": "special"},
        results={"送料無料": True},
    ))

    out_path = tmp_path / "out.xlsx"
    patch_write(edited, SHOP_POLICY, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb["FreeShipping"]

    assert ws.max_row == 12
    assert [ws.cell(12, c).value for c in (2, 3, 4)] == ["プレミアム会員", "special", True]

    # 直前行（11行目）からスタイルがコピーされている
    for col in (2, 3, 4):
        assert _fonts_equal(ws.cell(12, col).font, ws.cell(11, col).font)
        assert ws.cell(12, col).number_format == ws.cell(11, col).number_format

    # 既存行は不変
    before_wb = openpyxl.load_workbook(SHOP_POLICY)
    before_ws = before_wb["FreeShipping"]
    for r in range(1, 12):
        assert [ws.cell(r, c).value for c in (2, 3, 4)] == \
               [before_ws.cell(r, c).value for c in (2, 3, 4)]


def test_patch_write_insert_rule_in_middle(tmp_path):
    from openl.patch_writer import patch_write

    edited = OpenLReader().read(SHOP_POLICY)
    table = edited.get_table("FreeShipping")
    new_rule = Rule(
        id=99,
        conditions={"会員種別": "一般会員", "購入金額": "1 .. 99"},
        results={"送料無料": False},
    )
    table.rules.insert(2, new_rule)  # rule1, rule2 の後に挿入

    out_path = tmp_path / "out.xlsx"
    patch_write(edited, SHOP_POLICY, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb["FreeShipping"]
    before_wb = openpyxl.load_workbook(SHOP_POLICY)
    before_ws = before_wb["FreeShipping"]

    assert ws.max_row == 12

    # 行3,4（rule1,rule2）は不変
    for r in (3, 4):
        assert [ws.cell(r, c).value for c in (2, 3, 4)] == \
               [before_ws.cell(r, c).value for c in (2, 3, 4)]

    # 行5 に新規ルールが挿入され、スタイルは行4からコピー
    assert [ws.cell(5, c).value for c in (2, 3, 4)] == ["一般会員", "1 .. 99", False]
    for col in (2, 3, 4):
        assert _fonts_equal(ws.cell(5, col).font, ws.cell(4, col).font)

    # 行6〜12 に元rule3〜rule9がシフトしている
    for offset, r in enumerate(range(6, 13)):
        orig_r = 5 + offset
        assert [ws.cell(r, c).value for c in (2, 3, 4)] == \
               [before_ws.cell(orig_r, c).value for c in (2, 3, 4)]


def test_patch_write_delete_rule(tmp_path):
    from openl.patch_writer import patch_write

    edited = OpenLReader().read(SHOP_POLICY)
    table = edited.get_table("FreeShipping")
    del table.rules[4]  # rule id=5 (一般会員 / 3000..9999) を削除

    out_path = tmp_path / "out.xlsx"
    patch_write(edited, SHOP_POLICY, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb["FreeShipping"]
    before_wb = openpyxl.load_workbook(SHOP_POLICY)
    before_ws = before_wb["FreeShipping"]

    assert ws.max_row == 10

    # 行3〜6（rule1〜4）は不変
    for r in range(3, 7):
        assert [ws.cell(r, c).value for c in (2, 3, 4)] == \
               [before_ws.cell(r, c).value for c in (2, 3, 4)]

    # 行7〜10 に元rule6〜rule9がシフトしている
    for offset, r in enumerate(range(7, 11)):
        orig_r = 8 + offset
        assert [ws.cell(r, c).value for c in (2, 3, 4)] == \
               [before_ws.cell(orig_r, c).value for c in (2, 3, 4)]


def test_patch_write_add_data_table_row(tmp_path):
    from openl.patch_writer import patch_write

    edited = OpenLReader().read(AUTO_POLICY)
    table = next(
        t for t in edited.tables
        if t.sheet_name == "Vocabulary" and getattr(t, "table_type", None) == "TheftRating"
    )
    table.rows.append(DataTableRow(data={"value": "VeryHigh"}))

    out_path = tmp_path / "out.xlsx"
    patch_write(edited, AUTO_POLICY, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb["Vocabulary"]
    before_wb = openpyxl.load_workbook(AUTO_POLICY)
    before_ws = before_wb["Vocabulary"]

    # 新しい行が9行目（直前の8行目の直後）に挿入される
    assert ws.cell(9, 3).value == "VeryHigh"
    assert _fonts_equal(ws.cell(9, 3).font, ws.cell(8, 3).font)

    # 既存の TheftRating 行 (6-8) と、後続の Datatype 宣言行は1行下にシフト
    for r in (6, 7, 8):
        assert ws.cell(r, 3).value == before_ws.cell(r, 3).value
    assert ws.cell(12, 3).value == before_ws.cell(11, 3).value  # "Datatype InjuryRating <String>"


def test_patch_write_add_spreadsheet_step(tmp_path):
    from openl.patch_writer import patch_write

    edited = OpenLReader().read(SHOP_POLICY)
    table = edited.get_table("Calculation")
    table.steps.append(SpreadsheetStep(label="Test", value="= 1 + 1", unit="テスト"))

    out_path = tmp_path / "out.xlsx"
    patch_write(edited, SHOP_POLICY, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb["Calculation"]

    assert ws.max_row == 6
    assert ws.cell(6, 2).value == "Test"
    assert ws.cell(6, 3).value == "テスト"
    assert ws.cell(6, 4).value == "= 1 + 1"
    assert ws.cell(6, 4).data_type == "s"
    for col in (2, 3, 4):
        assert _fonts_equal(ws.cell(6, col).font, ws.cell(5, col).font)


def test_patch_write_new_table_new_sheet_and_deleted_table(tmp_path):
    from openl.patch_writer import patch_write

    edited = OpenLReader().read(SHOP_POLICY)

    # (a) 既存シート(FreeShipping)に新規テーブルを追記
    edited.tables.append(SimpleDecisionTable(
        sheet_name="FreeShipping",
        title="",
        method_signature="SimpleRules Boolean IsExpressShipping (String memberType)",
        table_name="IsExpressShipping",
        conditions=[ColumnDef(name="会員種別", col_type="String", role="condition")],
        results=[ColumnDef(name="速達無料", col_type="String", role="result")],
        rules=[Rule(id=1, conditions={"会員種別": "プレミアム会員"}, results={"速達無料": True})],
        start_col=1,
    ))

    # (b) 新規シート(Promotions)に新規テーブルを追記
    edited.tables.append(SimpleDecisionTable(
        sheet_name="Promotions",
        title="",
        method_signature="SimpleRules Boolean IsPromotionEligible (String memberType)",
        table_name="IsPromotionEligible",
        conditions=[ColumnDef(name="会員種別", col_type="String", role="condition")],
        results=[ColumnDef(name="対象", col_type="String", role="result")],
        rules=[Rule(id=1, conditions={"会員種別": "一般会員"}, results={"対象": True})],
        start_col=1,
    ))

    # (c) CampaignTarget のテーブルを削除
    edited.tables = [t for t in edited.tables if t.sheet_name != "CampaignTarget"]

    out_path = tmp_path / "out.xlsx"
    patch_write(edited, SHOP_POLICY, out_path)

    wb = openpyxl.load_workbook(out_path)

    # (a) FreeShipping: 元の11行 + 空行1 + 新規テーブル3行 = 15行
    ws = wb["FreeShipping"]
    assert ws.max_row == 15
    assert ws.cell(12, 2).value is None  # 区切りの空行
    assert ws.cell(13, 2).value == "SimpleRules Boolean IsExpressShipping (String memberType)"
    assert [ws.cell(14, c).value for c in (2, 3)] == ["会員種別", "速達無料"]
    assert [ws.cell(15, c).value for c in (2, 3)] == ["プレミアム会員", True]

    # (b) Promotions: 新規シート、区切り無しで3行
    assert "Promotions" in wb.sheetnames
    ws_promo = wb["Promotions"]
    assert ws_promo.cell(1, 2).value == "SimpleRules Boolean IsPromotionEligible (String memberType)"
    assert [ws_promo.cell(2, c).value for c in (2, 3)] == ["会員種別", "対象"]
    assert [ws_promo.cell(3, c).value for c in (2, 3)] == ["一般会員", True]

    # (c) CampaignTarget: テーブルが削除されている
    result = OpenLReader().read(out_path)
    assert result.get_table("CampaignTarget") is None

    # PointRate / Calculation は完全に不変
    before_wb = openpyxl.load_workbook(SHOP_POLICY)
    before = _all_cell_values(before_wb)
    after = _all_cell_values(wb)
    assert after["PointRate"] == before["PointRate"]
    assert after["Calculation"] == before["Calculation"]


def test_patch_write_append_spreadsheet_table_applies_header_styling(tmp_path):
    from openl.patch_writer import patch_write

    edited = OpenLReader().read(SHOP_POLICY)
    edited.tables.append(SpreadsheetTable(
        sheet_name="NewCalc",
        title="",
        description="Spreadsheet SpreadsheetResult NewCalc()",
        parameters=[],
        steps=[SpreadsheetStep(label="Step1", value="= 1 + 1", unit="計算結果")],
        column_names=["Step", "Description", "Value"],
        start_col=1,
        notes=[],
    ))

    out_path = tmp_path / "out.xlsx"
    patch_write(edited, SHOP_POLICY, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb["NewCalc"]

    # 行1: シグネチャー, 行2: ヘッダー(Step/Description/Value), 行3: ステップ
    assert ws.cell(1, 2).value == "Spreadsheet SpreadsheetResult NewCalc()"
    assert [ws.cell(2, c).value for c in (2, 3, 4)] == ["Step", "Description", "Value"]
    assert [ws.cell(3, c).value for c in (2, 3, 4)] == ["Step1", "計算結果", "= 1 + 1"]

    # ヘッダー行(2)はBold
    for col in (2, 3, 4):
        assert ws.cell(2, col).font.bold is True

    # 末尾ステップ行(3)もBold + 上部ボーダー
    for col in (2, 3, 4):
        assert ws.cell(3, col).font.bold is True
        assert ws.cell(3, col).border.top.style == "thin"


def test_patch_write_condition_column_added_recreates_table(tmp_path):
    from openl.patch_writer import patch_write

    edited = OpenLReader().read(SHOP_POLICY)
    table = edited.get_table("FreeShipping")
    table.conditions.append(ColumnDef(name="新条件", col_type="String", role="condition"))
    for rule in table.rules:
        rule.conditions["新条件"] = "test"

    out_path = tmp_path / "out.xlsx"
    patch_write(edited, SHOP_POLICY, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb["FreeShipping"]

    # 行数は変わらない（9ルール + シグネチャ行 + ヘッダー行 = 11行）
    assert ws.max_row == 11
    assert [ws.cell(2, c).value for c in (2, 3, 4, 5)] == ["会員種別", "購入金額", "新条件", "送料無料"]
    assert ws.cell(3, 4).value == "test"
    assert ws.cell(11, 4).value == "test"

    # 再パースすると新しい列構成で読み取れる
    result = OpenLReader().read(out_path)
    fs = result.get_table("FreeShipping")
    assert [c.name for c in fs.conditions] == ["会員種別", "購入金額", "新条件"]
    assert len(fs.rules) == 9
    assert fs.rules[0].conditions["新条件"] == "test"

    # 他のシートは完全に不変
    before_wb = openpyxl.load_workbook(SHOP_POLICY)
    before = _all_cell_values(before_wb)
    after = _all_cell_values(wb)
    for sheet in ("PointRate", "CampaignTarget", "Calculation"):
        assert after[sheet] == before[sheet]


def test_patch_write_recreate_table_preserves_trailing_separator_rows(tmp_path):
    from openl.patch_writer import patch_write
    from openl.reader import read_with_positions

    edited = OpenLReader().read(AUTO_POLICY)
    table = next(
        t for t in edited.tables
        if t.table_kind == "SimpleDecisionTable" and t.table_name == "VehicleTheftRating"
    )
    table.conditions.append(ColumnDef(name="新条件", col_type="String", role="condition"))
    for rule in table.rules:
        rule.conditions["新条件"] = "test"

    out_path = tmp_path / "out.xlsx"
    patch_write(edited, AUTO_POLICY, out_path)

    before_wb = openpyxl.load_workbook(AUTO_POLICY)
    after_wb = openpyxl.load_workbook(out_path)
    before_ws = before_wb["Vehicle-Eligibility"]
    after_ws = after_wb["Vehicle-Eligibility"]

    # VehicleTheftRating は元の位置(4-10)のまま、新しい列が追加される
    assert [after_ws.cell(5, c).value for c in (3, 4, 5, 6, 7)] == \
        ["Body Type", "Price", "High Theft List", "新条件", "Theft Rating"]
    assert after_ws.cell(6, 6).value == "test"

    # 区切りの空行 (11-15) はそのまま残る
    for r in range(11, 16):
        assert all(c.value is None for c in after_ws[r])

    # 2番目のテーブル(VehicleInjuryRating)は行位置・内容ともに不変
    # （末尾の None セルはシートの max_column 再計算により変動するため比較対象外）
    for r in range(16, 28):
        before_vals = [c.value for c in before_ws[r]]
        after_vals = [c.value for c in after_ws[r]]
        while before_vals and before_vals[-1] is None:
            before_vals.pop()
        while after_vals and after_vals[-1] is None:
            after_vals.pop()
        assert after_vals == before_vals

    # 再パースしても2番目のテーブルは元のstart_rowで見つかる
    result = read_with_positions(out_path)
    injury = next(
        p for p in result
        if p.table.sheet_name == "Vehicle-Eligibility" and p.table.table_kind == "SimpleDecisionTable"
        and p.table.table_name == "VehicleInjuryRating"
    )
    assert injury.start_row == 16


def test_patch_write_recreate_spreadsheet_table_respects_start_col(tmp_path):
    from openl.patch_writer import patch_write
    from openl.reader import read_with_positions

    edited = OpenLReader().read(AUTO_POLICY)
    table = next(
        t for t in edited.tables
        if t.sheet_name == "Calculation" and t.table_kind == "SpreadsheetTable"
        and _table_identity(t)[2] == "DetermineVehiclePremium"
    )
    assert table.start_col == 2
    table.column_names = table.column_names + ["Unit"]

    out_path = tmp_path / "out.xlsx"
    patch_write(edited, AUTO_POLICY, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb["Calculation"]

    # start_col=2 → シグネチャー・ヘッダーは列C(3列目)から始まる
    assert ws.cell(9, 2).value is None
    assert ws.cell(9, 3).value == table.description
    assert [ws.cell(10, c).value for c in (3, 4, 5, 6)] == ["Step", "Description", "Value", "Unit"]
    assert ws.cell(10, 2).value is None

    # 後続テーブル(DetermineDriverPremium)の位置は不変
    result = read_with_positions(out_path)
    driver_premium = next(
        p for p in result
        if p.table.sheet_name == "Calculation" and p.table.table_kind == "SpreadsheetTable"
        and _table_identity(p.table)[2] == "DetermineDriverPremium"
    )
    assert driver_premium.start_row == 29

    # 再パースしても start_col=2 を維持
    recreated = next(
        p for p in result
        if p.table.sheet_name == "Calculation" and p.table.table_kind == "SpreadsheetTable"
        and _table_identity(p.table)[2] == "DetermineVehiclePremium"
    )
    assert recreated.table.start_col == 2


def test_patch_write_append_two_new_tables_to_new_sheet_inserts_separator(tmp_path):
    from openl.patch_writer import patch_write

    edited = OpenLReader().read(SHOP_POLICY)

    edited.tables.append(SimpleDecisionTable(
        sheet_name="NewSheet",
        title="",
        method_signature="SimpleRules Boolean TableA (String memberType)",
        table_name="TableA",
        conditions=[ColumnDef(name="会員種別", col_type="String", role="condition")],
        results=[ColumnDef(name="結果A", col_type="String", role="result")],
        rules=[Rule(id=1, conditions={"会員種別": "プレミアム会員"}, results={"結果A": True})],
        start_col=1,
    ))
    edited.tables.append(SimpleDecisionTable(
        sheet_name="NewSheet",
        title="",
        method_signature="SimpleRules Boolean TableB (String memberType)",
        table_name="TableB",
        conditions=[ColumnDef(name="会員種別", col_type="String", role="condition")],
        results=[ColumnDef(name="結果B", col_type="String", role="result")],
        rules=[Rule(id=1, conditions={"会員種別": "一般会員"}, results={"結果B": False})],
        start_col=1,
    ))

    out_path = tmp_path / "out.xlsx"
    patch_write(edited, SHOP_POLICY, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb["NewSheet"]

    # TableA: 行1-3 (シグネチャ/ヘッダー/データ)
    assert ws.cell(1, 2).value == "SimpleRules Boolean TableA (String memberType)"
    assert [ws.cell(2, c).value for c in (2, 3)] == ["会員種別", "結果A"]
    assert [ws.cell(3, c).value for c in (2, 3)] == ["プレミアム会員", True]

    # 行4: 区切りの空行
    assert ws.cell(4, 2).value is None

    # TableB: 行5-7
    assert ws.cell(5, 2).value == "SimpleRules Boolean TableB (String memberType)"
    assert [ws.cell(6, c).value for c in (2, 3)] == ["会員種別", "結果B"]
    assert [ws.cell(7, c).value for c in (2, 3)] == ["一般会員", False]

    assert ws.max_row == 7


def test_patch_write_integration_auto_policy_accidents_condition(tmp_path):
    from openl.patch_writer import patch_write

    edited = OpenLReader().read(AUTO_POLICY)
    # 注意: AutoPolicyCalculation.xlsx には Vocabulary シートにも
    # table_name="DriverRisk" の DataTable (Datatype) が存在するため、
    # table_kind で SimpleDecisionTable に絞り込む。
    table = next(
        t for t in edited.tables
        if t.table_kind == "SimpleDecisionTable" and t.table_name == "DriverRisk"
    )
    rule = next(r for r in table.rules if r.id == 2)
    assert rule.conditions["Accidents"] == ">2"
    rule.conditions["Accidents"] = ">=10"

    out_path = tmp_path / "out.xlsx"
    patch_write(edited, AUTO_POLICY, out_path)

    before_wb = openpyxl.load_workbook(AUTO_POLICY)
    after_wb = openpyxl.load_workbook(out_path)
    before_ws = before_wb["Driver-Eligibility"]
    after_ws = after_wb["Driver-Eligibility"]

    # (a) 値が正しく変わる
    assert before_ws.cell(28, 4).value == ">2"
    assert after_ws.cell(28, 4).value == ">=10"

    # (b) ヘッダー行のBoldフォント・テーマカラー（背景塗りつぶし）が保持される
    before_header = before_ws.cell(26, 4)
    after_header = after_ws.cell(26, 4)
    assert after_header.font.bold is True and before_header.font.bold is True
    assert after_header.fill.fgColor.theme == before_header.fill.fgColor.theme
    assert after_header.fill.fgColor.tint == before_header.fill.fgColor.tint

    # (c) 編集対象外の行数・列幅・行高は不変
    assert after_ws.max_row == before_ws.max_row == 31
    for r in range(25, 32):
        assert after_ws.row_dimensions[r].height == before_ws.row_dimensions[r].height
    for col in "ABCDEFG":
        assert after_ws.column_dimensions[col].width == before_ws.column_dimensions[col].width

    # 他のシートは完全に不変
    assert _all_cell_values(after_wb)["Vehicle-Eligibility"] == _all_cell_values(before_wb)["Vehicle-Eligibility"]
